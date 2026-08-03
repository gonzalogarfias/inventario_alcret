"""Vistas de auditoría.

Correcciones Fase 4:
  1. StreamingHttpResponse CSV corregido (generador real)
  2. Uso de .iterator() para memoria eficiente
  3. Logging de exports
"""

import csv
import io
import logging

from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.http import HttpResponse, StreamingHttpResponse
from django.views.generic import ListView
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from apps.shared.csv_utils import sanitizar_celda

from .models import AuditLog

logger = logging.getLogger(__name__)


class AuditLogListView(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    """Lista paginada de registros de auditoría.

    Requiere permiso 'usuarios.puede_ver_auditoria_completa'.
    """

    model = AuditLog
    template_name = "auditoria/auditlog_list.html"
    context_object_name = "logs"
    paginate_by = 30
    permission_required = "usuarios.puede_ver_auditoria_completa"

    def get_queryset(self):
        qs = AuditLog.objects.select_related("usuario").order_by("-timestamp")
        evento = self.request.GET.get("evento")
        q = self.request.GET.get("q")
        if evento:
            qs = qs.filter(evento=evento)
        if q:
            qs = qs.filter(usuario__email__icontains=q) | qs.filter(
                ip_address__icontains=q
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["eventos"] = AuditLog.Evento.choices
        return ctx


auditlog_list = AuditLogListView.as_view()


def _generar_csv_auditoria(evento=None):
    """Generador que yield filas de CSV para StreamingHttpResponse."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Evento", "Usuario", "IP", "Datos", "Fecha"])
    yield buffer.getvalue()
    buffer.seek(0)
    buffer.truncate(0)

    qs = AuditLog.objects.select_related("usuario").order_by("-timestamp")
    if evento:
        qs = qs.filter(evento=evento)

    for log in qs.iterator():
        writer.writerow([
            log.evento,
            sanitizar_celda(str(log.usuario or "")),
            log.ip_address,
            sanitizar_celda(str(log.datos)),
            str(log.timestamp),
        ])
        yield buffer.getvalue()
        buffer.seek(0)
        buffer.truncate(0)


@login_required
@permission_required("usuarios.puede_ver_auditoria_completa", raise_exception=True)
def exportar_auditoria_csv(request):
    """Exporta registros de auditoría a CSV con streaming."""
    evento = request.GET.get("evento")
    response = StreamingHttpResponse(
        _generar_csv_auditoria(evento),
        content_type="text/csv",
    )
    response["Content-Disposition"] = 'attachment; filename="auditoria.csv"'
    logger.info("Exportación CSV de auditoría por %s", request.user.email)
    return response


@login_required
@permission_required("usuarios.puede_ver_auditoria_completa", raise_exception=True)
def exportar_auditoria_excel(request):
    """Exporta registros de auditoría a Excel (.xlsx)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría"
    ws.append(["Evento", "Usuario", "IP", "Datos", "Fecha"])

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    qs = AuditLog.objects.select_related("usuario").order_by("-timestamp")
    evento = request.GET.get("evento")
    if evento:
        qs = qs.filter(evento=evento)

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    if start_date:
        qs = qs.filter(timestamp__date__gte=start_date)
    if end_date:
        qs = qs.filter(timestamp__date__lte=end_date)

    for log in qs.iterator():
        ws.append([
            log.evento,
            sanitizar_celda(str(log.usuario or "")),
            log.ip_address,
            sanitizar_celda(str(log.datos)),
            str(log.timestamp),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="auditoria.xlsx"'
    wb.save(response)
    logger.info("Exportación Excel de auditoría por %s", request.user.email)
    return response
