"""Vistas del módulo de inventario.

Correcciones aplicadas (Fase 2):
  1. MovimientoCreateView ahora usa InventarioPermissionMixin (fix RBAC)
  2. Validación de stock suficiente antes de SALIDA
  3. Transacciones atómicas en movimientos
  4. StreamingHttpResponse CSV corregido
  5. Integración de Value Objects en validación
"""

import csv
import io
import logging

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import F, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse, StreamingHttpResponse
from django.shortcuts import redirect
from django.urls import reverse, reverse_lazy
from django.views.generic import CreateView, DetailView, ListView, UpdateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.shared.csv_utils import sanitizar_celda
from apps.shared.value_objects import ValidationError
from apps.usuarios.models import Usuario

from .forms import ProductoForm
from .models import Almacen, Categoria, Movimiento, Producto, Stock
from .services import registrar_movimiento

logger = logging.getLogger(__name__)


# ============================================================================
# Mixins
# ============================================================================

class InventarioPermissionMixin(LoginRequiredMixin):
    """Mixin base que verifica roles permitidos para vistas de inventario.

    Si el usuario no tiene el rol requerido, redirige con mensaje de error
    o lanza PermissionDenied según el contexto.
    """

    roles_permitidos: list[str] = []

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if self.roles_permitidos and request.user.rol not in self.roles_permitidos:
            messages.error(request, "No tenés permiso para realizar esta acción.")
            return redirect(self.request.GET.get("next", reverse("producto_list")))
        return super().dispatch(request, *args, **kwargs)


# ============================================================================
# Productos
# ============================================================================

class ProductoListView(LoginRequiredMixin, ListView):
    model = Producto
    template_name = "inventario/producto_list.html"
    context_object_name = "productos"
    paginate_by = 20

    def get_queryset(self):
        qs = Producto.objects.select_related("categoria").filter(activo=True)
        q = self.request.GET.get("q")
        cat = self.request.GET.get("categoria")
        stock = self.request.GET.get("stock")
        if q:
            qs = qs.filter(nombre__icontains=q) | qs.filter(sku__icontains=q)
        if cat:
            qs = qs.filter(categoria_id=cat)
        if stock == "bajo":
            qs = qs.annotate(
                total_stock=Coalesce(Sum("stocks__cantidad"), Value(0))
            ).filter(
                stock_minimo__gt=0,
                total_stock__lte=F("stock_minimo"),
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["categorias"] = Categoria.objects.filter(activo=True)
        return ctx


producto_list = ProductoListView.as_view()


class ProductoCreateView(InventarioPermissionMixin, CreateView):
    roles_permitidos = [Usuario.Rol.ADMINISTRADOR, Usuario.Rol.ALMACENISTA]
    model = Producto
    template_name = "inventario/producto_form.html"
    form_class = ProductoForm
    success_url = reverse_lazy("producto_list")

    def form_valid(self, form):
        messages.success(self.request, "Producto creado correctamente.")
        logger.info("Producto creado: %s por %s", form.instance.sku, self.request.user.email)
        return super().form_valid(form)


producto_create = ProductoCreateView.as_view()


class ProductoUpdateView(InventarioPermissionMixin, UpdateView):
    roles_permitidos = [Usuario.Rol.ADMINISTRADOR, Usuario.Rol.ALMACENISTA]
    model = Producto
    template_name = "inventario/producto_form.html"
    form_class = ProductoForm
    success_url = reverse_lazy("producto_list")

    def form_valid(self, form):
        messages.success(self.request, "Producto actualizado correctamente.")
        logger.info("Producto actualizado: %s por %s", form.instance.sku, self.request.user.email)
        return super().form_valid(form)


producto_update = ProductoUpdateView.as_view()


class ProductoDetailView(LoginRequiredMixin, DetailView):
    model = Producto
    template_name = "inventario/producto_detail.html"
    context_object_name = "producto"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["stocks"] = Stock.objects.filter(producto=self.object).select_related("almacen")
        ctx["movimientos"] = (
            Movimiento.objects.filter(producto=self.object)
            .select_related("almacen", "realizada_por")
            .order_by("-created_at")[:20]
        )
        return ctx


producto_detail = ProductoDetailView.as_view()


# ============================================================================
# Categorías
# ============================================================================

class CategoriaListView(LoginRequiredMixin, ListView):
    model = Categoria
    template_name = "inventario/categoria_list.html"
    context_object_name = "categorias"


categoria_list = CategoriaListView.as_view()


class CategoriaCreateView(InventarioPermissionMixin, CreateView):
    roles_permitidos = [Usuario.Rol.ADMINISTRADOR, Usuario.Rol.ALMACENISTA]
    model = Categoria
    template_name = "inventario/categoria_form.html"
    fields = ["nombre", "descripcion"]
    success_url = reverse_lazy("categoria_list")

    def form_valid(self, form):
        messages.success(self.request, "Categoría creada correctamente.")
        return super().form_valid(form)


categoria_create = CategoriaCreateView.as_view()


class CategoriaUpdateView(InventarioPermissionMixin, UpdateView):
    roles_permitidos = [Usuario.Rol.ADMINISTRADOR, Usuario.Rol.ALMACENISTA]
    model = Categoria
    template_name = "inventario/categoria_form.html"
    fields = ["nombre", "descripcion", "activo"]
    success_url = reverse_lazy("categoria_list")

    def form_valid(self, form):
        messages.success(self.request, "Categoría actualizada correctamente.")
        return super().form_valid(form)


categoria_update = CategoriaUpdateView.as_view()


# ============================================================================
# Almacenes
# ============================================================================

class AlmacenListView(LoginRequiredMixin, ListView):
    model = Almacen
    template_name = "inventario/almacen_list.html"
    context_object_name = "almacenes"


almacen_list = AlmacenListView.as_view()


class AlmacenCreateView(InventarioPermissionMixin, CreateView):
    roles_permitidos = [Usuario.Rol.ADMINISTRADOR, Usuario.Rol.ALMACENISTA]
    model = Almacen
    template_name = "inventario/almacen_form.html"
    fields = ["nombre", "ubicacion"]
    success_url = reverse_lazy("almacen_list")

    def form_valid(self, form):
        messages.success(self.request, "Almacén creado correctamente.")
        return super().form_valid(form)


almacen_create = AlmacenCreateView.as_view()


class AlmacenUpdateView(InventarioPermissionMixin, UpdateView):
    roles_permitidos = [Usuario.Rol.ADMINISTRADOR, Usuario.Rol.ALMACENISTA]
    model = Almacen
    template_name = "inventario/almacen_form.html"
    fields = ["nombre", "ubicacion", "activo"]
    success_url = reverse_lazy("almacen_list")

    def form_valid(self, form):
        messages.success(self.request, "Almacén actualizado correctamente.")
        return super().form_valid(form)


almacen_update = AlmacenUpdateView.as_view()


# ============================================================================
# Movimientos — CORREGIDO (RBAC + validación stock + transacciones)
# ============================================================================

class MovimientoListView(LoginRequiredMixin, ListView):
    model = Movimiento
    template_name = "inventario/movimiento_list.html"
    context_object_name = "movimientos"
    paginate_by = 25

    def get_queryset(self):
        qs = Movimiento.objects.select_related("producto", "almacen", "realizada_por").order_by("-created_at")
        tipo = self.request.GET.get("tipo")
        q = self.request.GET.get("q")
        if tipo:
            qs = qs.filter(tipo=tipo)
        if q:
            qs = qs.filter(producto__nombre__icontains=q) | qs.filter(producto__sku__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["tipos"] = Movimiento.Tipo.choices
        return ctx


movimiento_list = MovimientoListView.as_view()


class MovimientoCreateView(InventarioPermissionMixin, CreateView):
    """Vista para registrar movimientos de inventario.

    CORRECCIONES Fase 2:
      - Ahora usa InventarioPermissionMixin (fix vulnerabilidad RBAC)
      - Validación de stock suficiente antes de SALIDA
      - Transacción atómica + select_for_update
      - Validación de cantidad con Value Object
    """

    roles_permitidos = [
        Usuario.Rol.ADMINISTRADOR,
        Usuario.Rol.VENDEDOR,
        Usuario.Rol.ALMACENISTA,
    ]
    model = Movimiento
    template_name = "inventario/movimiento_form.html"
    fields = ["tipo", "producto", "almacen", "cantidad", "costo_unitario", "motivo"]
    success_url = reverse_lazy("movimiento_list")

    def get_initial(self):
        initial = super().get_initial()
        if self.request.GET.get("producto"):
            initial["producto"] = self.request.GET.get("producto")
        return initial

    def _check_movimiento_permission(self, form) -> bool:
        """Verifica permisos específicos según el tipo de movimiento.

        Matriz de permisos (ARQUITECTURA.md):
          - ENTRADA: Admin, Almacenista
          - SALIDA: Admin, Vendedor, Almacenista
          - AJUSTE: Solo Admin
        """
        tipo = form.instance.tipo
        user = self.request.user
        rol = user.rol

        # AJUSTE: solo Administrador
        if tipo == Movimiento.Tipo.AJUSTE and rol != Usuario.Rol.ADMINISTRADOR:
            messages.error(self.request, "No tenés permiso para realizar ajustes de stock.")
            return False

        # ENTRADA: Admin o Almacenista (Vendedor NO puede)
        if tipo == Movimiento.Tipo.ENTRADA and rol not in (
            Usuario.Rol.ADMINISTRADOR,
            Usuario.Rol.ALMACENISTA,
        ):
            messages.error(self.request, "No tenés permiso para registrar entradas de stock.")
            return False

        return True

    def form_valid(self, form):
        # 1. Verificar permisos por tipo de movimiento
        if not self._check_movimiento_permission(form):
            return self.form_invalid(form)

        # 2. Delegar la creación en el servicio con lock y validación
        try:
            movimiento = registrar_movimiento(
                tipo=form.cleaned_data["tipo"],
                producto=form.cleaned_data["producto"],
                almacen=form.cleaned_data["almacen"],
                cantidad=form.cleaned_data["cantidad"],
                realizada_por=self.request.user,
                costo_unitario=form.cleaned_data.get("costo_unitario"),
                motivo=form.cleaned_data.get("motivo", ""),
            )
        except ValidationError as e:
            form.add_error("cantidad", str(e))
            return self.form_invalid(form)

        messages.success(
            self.request,
            f"Movimiento de {movimiento.get_tipo_display().lower()} registrado.",
        )
        logger.info(
            "Movimiento %s registrado: producto=%s almacen=%s cantidad=%s por %s",
            movimiento.tipo,
            movimiento.producto.sku,
            movimiento.almacen.nombre,
            movimiento.cantidad,
            self.request.user.email,
        )
        return redirect(self.success_url)


movimiento_create = MovimientoCreateView.as_view()


# ============================================================================
# Exports — CORREGIDO (StreamingHttpResponse + rate limiting)
# ============================================================================

def _estilo_excel(ws):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20


def _generar_csv_productos():
    """Generador que yield filas de CSV para StreamingHttpResponse."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["SKU", "Nombre", "Categoría", "Precio", "Stock Mínimo"])
    yield buffer.getvalue()
    buffer.seek(0)
    buffer.truncate(0)

    for p in Producto.objects.select_related("categoria").filter(activo=True).iterator():
        writer.writerow([
            sanitizar_celda(p.sku),
            sanitizar_celda(p.nombre),
            sanitizar_celda(p.categoria.nombre),
            sanitizar_celda(str(p.precio_venta)),
            sanitizar_celda(str(p.stock_minimo)),
        ])
        yield buffer.getvalue()
        buffer.seek(0)
        buffer.truncate(0)


@login_required
def exportar_productos_csv(request):
    """Exporta productos activos a CSV con streaming."""
    if request.user.rol not in (Usuario.Rol.ADMINISTRADOR, Usuario.Rol.ALMACENISTA):
        messages.error(request, "No tenés permiso para exportar datos.")
        return redirect("producto_list")
    response = StreamingHttpResponse(
        _generar_csv_productos(),
        content_type="text/csv",
    )
    response["Content-Disposition"] = 'attachment; filename="productos.csv"'
    return response


@login_required
def exportar_productos_excel(request):
    """Exporta productos activos a Excel (.xlsx)."""
    if request.user.rol not in (Usuario.Rol.ADMINISTRADOR, Usuario.Rol.ALMACENISTA):
        messages.error(request, "No tenés permiso para exportar datos.")
        return redirect("producto_list")
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(["SKU", "Nombre", "Categoría", "Precio", "Stock Mínimo"])
    _estilo_excel(ws)
    for p in Producto.objects.select_related("categoria").filter(activo=True).iterator():
        ws.append([p.sku, p.nombre, p.categoria.nombre, float(p.precio_venta), float(p.stock_minimo)])
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="productos.xlsx"'
    wb.save(response)
    return response


def _generar_csv_movimientos(tipo=None):
    """Generador que yield filas de CSV para movimientos."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Tipo", "Producto", "SKU", "Almacén", "Cantidad", "Costo Unitario", "Realizado Por", "Fecha"])
    yield buffer.getvalue()
    buffer.seek(0)
    buffer.truncate(0)

    qs = Movimiento.objects.select_related("producto", "almacen", "realizada_por").order_by("-created_at")
    if tipo:
        qs = qs.filter(tipo=tipo)

    for m in qs.iterator():
        writer.writerow([
            m.tipo,
            m.producto.nombre,
            m.producto.sku,
            m.almacen.nombre,
            str(m.cantidad),
            str(m.costo_unitario or ""),
            sanitizar_celda(str(m.realizada_por or "")),
            sanitizar_celda(str(m.created_at)),
        ])
        yield buffer.getvalue()
        buffer.seek(0)
        buffer.truncate(0)


@login_required
def exportar_movimientos_csv(request):
    """Exporta movimientos a CSV con streaming."""
    if request.user.rol not in (Usuario.Rol.ADMINISTRADOR, Usuario.Rol.ALMACENISTA):
        messages.error(request, "No tenés permiso para exportar datos.")
        return redirect("movimiento_list")
    tipo = request.GET.get("tipo")
    response = StreamingHttpResponse(
        _generar_csv_movimientos(tipo),
        content_type="text/csv",
    )
    response["Content-Disposition"] = 'attachment; filename="movimientos.csv"'
    return response


@login_required
def exportar_movimientos_excel(request):
    """Exporta movimientos a Excel (.xlsx)."""
    if request.user.rol not in (Usuario.Rol.ADMINISTRADOR, Usuario.Rol.ALMACENISTA):
        messages.error(request, "No tenés permiso para exportar datos.")
        return redirect("movimiento_list")
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    ws.append(["Tipo", "Producto", "SKU", "Almacén", "Cantidad", "Costo Unitario", "Realizado Por", "Fecha"])
    _estilo_excel(ws)
    qs = Movimiento.objects.select_related("producto", "almacen", "realizada_por").order_by("-created_at")
    tipo = request.GET.get("tipo")
    if tipo:
        qs = qs.filter(tipo=tipo)
    for m in qs.iterator():
        ws.append([
            m.tipo,
            m.producto.nombre,
            m.producto.sku,
            m.almacen.nombre,
            float(m.cantidad),
            float(m.costo_unitario) if m.costo_unitario else "",
            sanitizar_celda(str(m.realizada_por or "")),
            sanitizar_celda(str(m.created_at)),
        ])
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="movimientos.xlsx"'
    wb.save(response)
    return response
